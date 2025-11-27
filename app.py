import streamlit as st
import pandas as pd
import numpy as np
import re
import json
import math
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(page_title="Tender Pricing App", layout="wide")

# ---------- Defaults meta (for colouring old vs new groups) ----------
DEFAULT_GROUP_NAMES = {
    "CORFLUTE_3MM",
    "SCREENBOARD_2MM",
    "POSTER_BOARD",
    "WINDOW_SUPERCLING",
    "BANNER_SYNTHETIC",
    "FERROUS",
    "VINYL_MPI1105",
    "VINYL_3M7725",
    "VINYL_ARLON8000",
    "BACKLIT_DURATRAN",
    "BRAILLE_SIGNS",
}

# ---------- Helpers ----------

def num_to_col_letters(n: int) -> str:
    """1 -> A, 2 -> B, ... 27 -> AA, etc."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def to_excel_view(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build a view that looks like Excel:
    - Column headers: A, B, C...
    - Row numbers: 1, 2, 3...
    - Row 1 contains the original header names.
    """
    letters = [num_to_col_letters(i + 1) for i in range(len(df.columns))]
    data = [list(df.columns)] + df.astype(object).values.tolist()
    excel_df = pd.DataFrame(data, columns=letters)
    excel_df.index = range(1, len(excel_df) + 1)
    excel_df.index.name = ""
    return excel_df


def parse_dimension_to_sqm(dim_str: str) -> float:
    """
    Parse strings like '841mm x 1189mm', '594 x 841mm', '1.2m x 2m' to sqm.
    Assumptions:
    - mm, cm, m supported
    - if no unit, assume mm
    """
    if pd.isna(dim_str):
        return np.nan

    s = str(dim_str).lower()
    s = s.replace("×", "x")

    # Find two numeric values with optional unit
    matches = re.findall(r"(\d+(\.\d+)?)\s*(mm|cm|m)?", s)
    if len(matches) < 2:
        return np.nan

    (v1, _, u1) = matches[0]
    (v2, _, u2) = matches[1]

    v1 = float(v1)
    v2 = float(v2)

    def to_m(v, u):
        if u == "cm":
            return v / 100.0
        if u == "m":
            return v
        # default or mm
        return v / 1000.0

    w = to_m(v1, u1)
    h = to_m(v2, u2)
    return w * h


def detect_side(text, ds_synonyms, ss_synonyms, default="SS"):
    """Return 'DS' or 'SS' based on synonyms found in text."""
    if pd.isna(text):
        return default

    s = str(text).strip().lower()
    if any(tok in s for tok in ds_synonyms):
        return "DS"
    if any(tok in s for tok in ss_synonyms):
        return "SS"
    return default


def build_items_from_rows(
    df,
    col_letters_map,
    size_col_letter,
    material_col_letter,
    qty_annum_col_letter,
    qty_run_col_letter,
    runs_pa_col_letter,
    side_mode,
    side_col_letter,
    side_source_letter,
    ds_synonyms,
    ss_synonyms,
):
    """
    Items are in rows (BP-style).
    Supports either:
    - explicit Qty per run column, OR
    - Qty per annum + Runs p.a. column -> Qty per run = Qty per annum / Runs p.a.
    """
    letter_to_header = col_letters_map
    result_rows = []

    size_col = letter_to_header.get(size_col_letter)
    mat_col = letter_to_header.get(material_col_letter) if material_col_letter else None
    qty_annum_col = letter_to_header.get(qty_annum_col_letter) if qty_annum_col_letter else None
    qty_run_col = letter_to_header.get(qty_run_col_letter) if qty_run_col_letter else None
    runs_pa_col = letter_to_header.get(runs_pa_col_letter) if runs_pa_col_letter else None

    side_col = (
        letter_to_header.get(side_col_letter)
        if side_mode == "Separate column" and side_col_letter
        else None
    )
    side_src_col = (
        letter_to_header.get(side_source_letter)
        if side_mode == "Embedded in another column" and side_source_letter
        else None
    )

    for idx, row in df.iterrows():
        size_val = row[size_col] if size_col else None
        material_val = row[mat_col] if mat_col else None

        qty_annum = pd.to_numeric(row[qty_annum_col], errors="coerce") if qty_annum_col else np.nan

        if qty_run_col:
            qty_run = pd.to_numeric(row[qty_run_col], errors="coerce")
        elif runs_pa_col and qty_annum_col:
            runs_pa = pd.to_numeric(row[runs_pa_col], errors="coerce")
            if not np.isnan(qty_annum) and not np.isnan(runs_pa) and runs_pa > 0:
                qty_run = qty_annum / runs_pa
            else:
                qty_run = np.nan
        else:
            qty_run = np.nan

        # Side detection
        if side_mode == "Separate column" and side_col:
            side_raw = row[side_col]
        elif side_mode == "Embedded in another column" and side_src_col:
            side_raw = row[side_src_col]
        else:
            side_raw = None

        side = detect_side(side_raw, ds_synonyms, ss_synonyms, default="SS")

        sqm_per_unit = parse_dimension_to_sqm(size_val)

        sqm_per_annum = (
            sqm_per_unit * qty_annum
            if (not np.isnan(sqm_per_unit) and not np.isnan(qty_annum))
            else np.nan
        )
        sqm_per_run = (
            sqm_per_unit * qty_run
            if (not np.isnan(sqm_per_unit) and not np.isnan(qty_run))
            else np.nan
        )

        result_rows.append(
            {
                "Source Row": idx + 2,  # Excel-style row: +2 because header is row 1
                "Size": size_val,
                "Material": material_val,
                "Qty per annum": qty_annum,
                "Qty per run": qty_run,
                "Side": side,
                "SQM per unit": sqm_per_unit,
                "SQM per annum": sqm_per_annum,
                "SQM per run": sqm_per_run,
            }
        )

    return pd.DataFrame(result_rows)


def build_items_from_columns(
    df,
    size_row,
    material_row,
    qty_annum_row,
    qty_run_row,
    runs_pa_row,
    side_mode,
    side_row,
    side_source_row,
    ds_synonyms,
    ss_synonyms,
):
    """
    Items are in columns (Foot Locker-style).
    Excel rows are used for mapping. Supports:
    - explicit Qty per run row, OR
    - Qty per annum row + Runs p.a. row -> Qty per run = Qty per annum / Runs p.a.
    """
    max_row, max_col = df.shape
    result_rows = []

    # Convert Excel row to df index (Excel row 2 -> df index 0)
    def excel_to_df_row(excel_row):
        return excel_row - 2

    size_r = excel_to_df_row(size_row) if size_row else None
    mat_r = excel_to_df_row(material_row) if material_row else None
    qty_annum_r = excel_to_df_row(qty_annum_row) if qty_annum_row else None
    qty_run_r = excel_to_df_row(qty_run_row) if qty_run_row else None
    runs_pa_r = excel_to_df_row(runs_pa_row) if runs_pa_row else None

    side_r = excel_to_df_row(side_row) if (side_mode == "Separate row" and side_row) else None
    side_src_r = (
        excel_to_df_row(side_source_row)
        if (side_mode == "Embedded in another row" and side_source_row)
        else None
    )

    for col_idx in range(max_col):
        col_letter = num_to_col_letters(col_idx + 1)

        size_val = df.iloc[size_r, col_idx] if size_r is not None else None
        material_val = df.iloc[mat_r, col_idx] if mat_r is not None else None

        qty_annum = (
            pd.to_numeric(df.iloc[qty_annum_r, col_idx], errors="coerce")
            if qty_annum_r is not None
            else np.nan
        )

        if qty_run_r is not None:
            qty_run = pd.to_numeric(df.iloc[qty_run_r, col_idx], errors="coerce")
        elif runs_pa_r is not None and qty_annum_r is not None:
            runs_pa = pd.to_numeric(df.iloc[runs_pa_r, col_idx], errors="coerce")
            if not np.isnan(qty_annum) and not np.isnan(runs_pa) and runs_pa > 0:
                qty_run = qty_annum / runs_pa
            else:
                qty_run = np.nan
        else:
            qty_run = np.nan

        # Skip totally empty items
        if (
            pd.isna(size_val)
            and pd.isna(material_val)
            and np.isnan(qty_annum)
            and np.isnan(qty_run)
        ):
            continue

        # Side detection
        if side_mode == "Separate row" and side_r is not None:
            side_raw = df.iloc[side_r, col_idx]
        elif side_mode == "Embedded in another row" and side_src_r is not None:
            side_raw = df.iloc[side_src_r, col_idx]
        else:
            side_raw = None

        side = detect_side(side_raw, ds_synonyms, ss_synonyms, default="SS")

        sqm_per_unit = parse_dimension_to_sqm(size_val)

        sqm_per_annum = (
            sqm_per_unit * qty_annum
            if (not np.isnan(sqm_per_unit) and not np.isnan(qty_annum))
            else np.nan
        )
        sqm_per_run = (
            sqm_per_unit * qty_run
            if (not np.isnan(sqm_per_unit) and not np.isnan(qty_run))
            else np.nan
        )

        result_rows.append(
            {
                "Source Column": col_letter,
                "Size": size_val,
                "Material": material_val,
                "Qty per annum": qty_annum,
                "Qty per run": qty_run,
                "Side": side,
                "SQM per unit": sqm_per_unit,
                "SQM per annum": sqm_per_annum,
                "SQM per run": sqm_per_run,
            }
        )

    return pd.DataFrame(result_rows)


# ---------- UI ----------

st.title("Tender Pricing App (Excel-style, Grouped Pricing)")

st.markdown(
    """
**Step 1:** Upload Excel and view all rows/columns (Excel-style A,B,C + 1,2,3)  
**Step 2:** Hide/Unhide rows & columns (without deleting)  
**Step 3:** Map fields (Size, Material, Qty, DS/SS) and calculate SQM + **Grouped Prices**
"""
)

# ---------- Initialise session_state ----------
if "group_assignments" not in st.session_state:
    st.session_state["group_assignments"] = {}
if "group_prices" not in st.session_state:
    st.session_state["group_prices"] = {}
if "group_volume_flags" not in st.session_state:
    st.session_state["group_volume_flags"] = {}  # NEW: group -> bool using volume tiers
if "material_overrides" not in st.session_state:
    st.session_state["material_overrides"] = {}  # kept for presets but unused in pricing
if "calc_df" not in st.session_state:
    st.session_state["calc_df"] = None
if "preset_file_loaded_once" not in st.session_state:
    st.session_state["preset_file_loaded_once"] = False
if "session_snapshots" not in st.session_state:
    st.session_state["session_snapshots"] = {}
if "extra_groups_to_show" not in st.session_state:
    st.session_state["extra_groups_to_show"] = []
if "reset_assign_widgets" not in st.session_state:
    st.session_state["reset_assign_widgets"] = False
# snapshot-able config defaults
if "ds_syn_input" not in st.session_state:
    st.session_state["ds_syn_input"] = "ds,double sided,double-sided,2s,2 sided,2sided,double"
if "ss_syn_input" not in st.session_state:
    st.session_state["ss_syn_input"] = "ss,single sided,single-sided,1s,1 sided,1sided,single"
if "double_sided_loading_percent" not in st.session_state:
    st.session_state["double_sided_loading_percent"] = 25.0
# hide/unhide state (columns/rows)
if "hidden_cols_letters" not in st.session_state:
    st.session_state["hidden_cols_letters"] = []
if "hidden_rows_numbers" not in st.session_state:
    st.session_state["hidden_rows_numbers"] = []
# volume tier defaults
if "tier_count" not in st.session_state:
    st.session_state["tier_count"] = 1
if "tier_thresholds" not in st.session_state:
    st.session_state["tier_thresholds"] = [250.0]
if "tier_discounts" not in st.session_state:
    st.session_state["tier_discounts"] = [0.0]

# Load default preset only once at very beginning (if nothing in state yet)
if not st.session_state["group_assignments"] and not st.session_state["group_prices"]:
    try:
        with open("material_groups_default.json", "r", encoding="utf-8") as f:
            preset = json.load(f)
        st.session_state["group_assignments"] = preset.get("group_assignments", {})
        st.session_state["group_prices"] = preset.get("group_prices", {})
        st.session_state["material_overrides"] = preset.get("material_overrides", {})
    except Exception:
        pass

uploaded_file = st.file_uploader(
    "Upload Excel file", type=["xlsx", "xls"], accept_multiple_files=False
)

if uploaded_file is None:
    st.info("Please upload an Excel file to begin.")
    st.stop()

# Use getvalue() so the file content is available on every rerun
file_bytes = uploaded_file.getvalue()

# --- Load sheet list ---
excel_file = pd.ExcelFile(BytesIO(file_bytes))
sheet_name = st.selectbox("Select sheet", options=excel_file.sheet_names)

# --- Read selected sheet into DataFrame ---
df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)

# Show the sheet in Excel-like view (A,B,C... and rows 1,2,3...)
st.subheader(f"Sheet preview (Excel-style): {sheet_name}")
excel_view = to_excel_view(df)
st.dataframe(excel_view)

# --- Build Excel-style column letter mapping (internal) ---
col_letters = {}
col_labels = {}
for i, col_name in enumerate(df.columns):
    letter = num_to_col_letters(i + 1)
    col_letters[letter] = col_name
    col_labels[letter] = f"{letter} - {col_name}"

# Helper to show dropdowns with friendly labels but keep track of the letter
def select_letter(label, options_letters, default_letter=None, key=None, allow_none=False, none_label="(none)"):
    options = []
    mapping = {}
    if allow_none:
        options.append(none_label)
        mapping[none_label] = None
    for ltr in options_letters:
        lab = col_labels[ltr]
        options.append(lab)
        mapping[lab] = ltr

    if default_letter is not None and default_letter in options_letters:
        default_label = col_labels[default_letter]
        default_index = options.index(default_label)
    else:
        default_index = 0

    choice = st.selectbox(label, options=options, index=default_index, key=key)
    return mapping[choice]

# ======================================================
# STEP 2: HIDE / UNHIDE COLUMNS & ROWS (for preview + export)
# ======================================================

st.header("Step 2 – Hide / Unhide Rows & Columns")

all_letters = list(col_letters.keys())

# Build options for columns
col_options_labels = [col_labels[ltr] for ltr in all_letters]

# Determine default labels based on saved letters
default_cols_labels = [
    col_labels[l] for l in st.session_state["hidden_cols_letters"] if l in col_labels
]

# Multiselect for columns to hide
cols_to_hide_labels = st.multiselect(
    "Select columns to HIDE (by Excel letter):",
    options=col_options_labels,
    default=default_cols_labels if "hidden_cols_multiselect" not in st.session_state else None,
    key="hidden_cols_multiselect",
)
# Convert labels back to letters
cols_to_hide_letters = [lab.split(" - ")[0] for lab in cols_to_hide_labels]
st.session_state["hidden_cols_letters"] = cols_to_hide_letters

# Rows to hide (Excel-style rows: include header row 1)
max_row = len(df) + 1  # +1 for header row
row_numbers = list(range(1, max_row + 1))

rows_to_hide_display = st.multiselect(
    "Select rows to HIDE (by Excel row number):",
    options=row_numbers,
    default=st.session_state["hidden_rows_numbers"] if "hidden_rows_multiselect" not in st.session_state else None,
    key="hidden_rows_multiselect",
)
st.session_state["hidden_rows_numbers"] = rows_to_hide_display

# Preview with hidden rows/cols (Excel-like view)
preview_excel_view = excel_view.copy()
if cols_to_hide_letters:
    preview_excel_view = preview_excel_view.drop(columns=cols_to_hide_letters)
if rows_to_hide_display:
    preview_excel_view = preview_excel_view.drop(index=rows_to_hide_display)

st.subheader(f"Preview with hidden rows/columns (Excel-style): {sheet_name}")
st.caption(
    "Preview hides selected rows/columns. Original workbook remains intact; "
    "exported file will mark them as hidden in Excel."
)
st.dataframe(preview_excel_view)

# Export with hidden rows/columns
st.subheader("Export with Hidden Rows / Columns")
if st.button("Prepare file with hidden rows/columns"):
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb[sheet_name]

    # Hide selected columns
    for letter in cols_to_hide_letters:
        ws.column_dimensions[letter].hidden = True

    # Hide selected rows (Excel rows directly)
    for r in rows_to_hide_display:
        ws.row_dimensions[r].hidden = True

    out_buf = BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    st.download_button(
        "Download workbook (with hidden rows/columns)",
        data=out_buf,
        file_name=f"{sheet_name}_hidden.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

# ======================================================
# STEP 3: SQM & PRICE CALCULATION
# ======================================================

st.header("Step 3 – SQM & Price Calculation")

st.markdown(
    """
Here you tell the app **where** the data lives (columns vs rows) and how DS/SS is encoded,
so it can calculate **square meters** and **pricing by material**.

In the **Material Pricing** area you can:
- Assign each material to a **Group name** (e.g. "3mm ACM", "Posters", "Window Vinyl").  
- Either **type a new group name** or **select an existing group** from a dropdown.  
- See **SQM totals per group** and **each group's base price per SQM**.  
- Configure **volume-based price breaks by SQM per annum**, using a base price per group and global % adjustments per SQM tier.  
- Opt **per group** whether to use those volume tiers or just the flat BASE price.
- **Old default groups show in green**, **new groups show in red**, and **default groups actually used in this Excel** show in a **bright highlight colour**.  
- Groups not used in this campaign are hidden by default but can be **unhidden from a dropdown**.  
- Save/Load **group presets** and **session snapshots** so you can reuse them next campaign/tender.
"""
)

layout_type = st.radio(
    "How are items laid out in this sheet?",
    ["Items are in rows (BP-style)", "Items are in columns (Foot Locker-style)"],
    key="layout_type_choice",
)

# DS/SS synonyms + loading
st.subheader("Double-sided / Single-sided configuration")

ds_syn_input = st.text_input(
    "Values meaning DOUBLE-SIDED (comma-separated)",
    value=st.session_state["ds_syn_input"],
    key="ds_syn_input",
)
ss_syn_input = st.text_input(
    "Values meaning SINGLE-SIDED (comma-separated)",
    value=st.session_state["ss_syn_input"],
    key="ss_syn_input",
)

ds_synonyms = [s.strip().lower() for s in ds_syn_input.split(",") if s.strip()]
ss_synonyms = [s.strip().lower() for s in ss_syn_input.split(",") if s.strip()]

double_sided_loading_percent = st.number_input(
    "Double-sided loading % (e.g. 25 for 25% extra over single-sided)",
    min_value=0.0,
    max_value=500.0,
    value=st.session_state["double_sided_loading_percent"],
    step=1.0,
    key="double_sided_loading_percent",
)

# Start from stored calculation (so it survives reruns)
calc_df = st.session_state["calc_df"]

if layout_type == "Items are in rows (BP-style)":
    st.subheader("Mapping (items in rows)")

    letters = list(col_letters.keys())

    # Try to auto-guess some defaults by header name
    headers_lower = {ltr: str(h).lower() for ltr, h in col_letters.items()}

    def guess_letter(substrings, fallback):
        for ltr, h in headers_lower.items():
            if any(sub in h for sub in substrings):
                return ltr
        return fallback

    size_default = guess_letter(["dim", "size"], letters[0] if letters else None)
    material_default = guess_letter(
        ["material", "stock", "substrate"], letters[0] if letters else None
    )
    qty_annum_default = guess_letter(
        ["annual", "per annum", "total annual volume", "pa"], letters[0] if letters else None
    )
    qty_run_default = guess_letter(
        ["per run", "run qty", "run quantity"], letters[0] if letters else None
    )
    runs_pa_default = guess_letter(
        ["runs p.a", "approx runs", "runs pa"], letters[0] if letters else None
    )

    size_col_letter = select_letter(
        "Size / Dimensions column",
        options_letters=letters,
        default_letter=size_default,
        key="size_col_letter_rows"
    )
    material_col_letter = select_letter(
        "Material name column",
        options_letters=letters,
        default_letter=material_default,
        key="material_col_letter_rows",
        allow_none=True,
    )
    qty_annum_col_letter = select_letter(
        "Quantity PER ANNUM column",
        options_letters=letters,
        default_letter=qty_annum_default,
        key="qty_annum_col_letter_rows",
        allow_none=True,
    )
    qty_run_col_letter = select_letter(
        "Quantity PER RUN column (optional — leave None to derive from runs p.a)",
        options_letters=letters,
        default_letter=qty_run_default,
        key="qty_run_col_letter_rows",
        allow_none=True,
    )
    runs_pa_col_letter = select_letter(
        "Runs PER ANNUM column (e.g. Approx runs p.a, optional)",
        options_letters=letters,
        default_letter=runs_pa_default,
        key="runs_pa_col_letter_rows",
        allow_none=True,
    )

    st.markdown("**Where is Single / Double-sided information?**")
    side_mode = st.selectbox(
        "Choose how DS/SS is stored:",
        ["Separate column", "Embedded in another column", "Not available (assume SS)"],
        key="side_mode_rows",
    )

    side_col_letter = None
    side_source_letter = None

    if side_mode == "Separate column":
        side_col_letter = select_letter(
            "Column that contains DS/SS values",
            options_letters=letters,
            key="side_col_letter_rows"
        )
    elif side_mode == "Embedded in another column":
        side_source_letter = select_letter(
            "Column where DS/SS text appears (e.g. Size or Description)",
            options_letters=letters,
            default_letter=size_col_letter,
            key="side_source_letter_rows"
        )

    if st.button("Calculate SQM & build item table", key="calc_rows"):
        calc_df = build_items_from_rows(
            df=df,
            col_letters_map=col_letters,
            size_col_letter=size_col_letter,
            material_col_letter=material_col_letter,
            qty_annum_col_letter=qty_annum_col_letter,
            qty_run_col_letter=qty_run_col_letter,
            runs_pa_col_letter=runs_pa_col_letter,
            side_mode=side_mode,
            side_col_letter=side_col_letter,
            side_source_letter=side_source_letter,
            ds_synonyms=ds_synonyms,
            ss_synonyms=ss_synonyms,
        )
        st.session_state["calc_df"] = calc_df

elif layout_type == "Items are in columns (Foot Locker-style)":
    st.subheader("Mapping (items in columns)")

    max_row, max_col = df.shape
    # Excel rows: header row is 1, df data starts at Excel row 2
    row_options = list(range(2, max_row + 2))

    size_row = st.selectbox(
        "Excel row that contains Size / Dimensions (across columns)",
        options=row_options,
        index=0,
        key="size_row_cols",
    )
    material_row = st.selectbox(
        "Excel row that contains Material name (across columns)",
        options=["(none)"] + row_options,
        index=0,
        key="material_row_cols",
    )
    qty_annum_row = st.selectbox(
        "Excel row that contains Quantity PER ANNUM (across columns)",
        options=["(none)"] + row_options,
        index=0,
        key="qty_annum_row_cols",
    )
    qty_run_row = st.selectbox(
        "Excel row that contains Quantity PER RUN (across columns)",
        options=["(none)"] + row_options,
        index=0,
        key="qty_run_row_cols",
    )
    runs_pa_row = st.selectbox(
        "Excel row that contains Runs PER ANNUM (across columns, optional)",
        options=["(none)"] + row_options,
        index=0,
        key="runs_pa_row_cols",
    )

    # Convert "(none)" to None
    material_row = None if material_row == "(none)" else material_row
    qty_annum_row = None if qty_annum_row == "(none)" else qty_annum_row
    qty_run_row = None if qty_run_row == "(none)" else qty_run_row
    runs_pa_row = None if runs_pa_row == "(none)" else runs_pa_row

    st.markdown("**Where is Single / Double-sided information?**")
    side_mode = st.selectbox(
        "Choose how DS/SS is stored:",
        ["Separate row", "Embedded in another row", "Not available (assume SS)"],
        key="side_mode_cols",
    )

    side_row = None
    side_source_row = None

    if side_mode == "Separate row":
        side_row = st.selectbox(
            "Excel row that contains DS/SS values (across columns)",
            options=row_options,
            key="side_row_cols",
        )
    elif side_mode == "Embedded in another row":
        side_source_row = st.selectbox(
            "Excel row where DS/SS text appears (e.g. in Size or Description row)",
            options=row_options,
            index=row_options.index(size_row) if size_row in row_options else 0,
            key="side_source_row_cols",
        )

    if st.button("Calculate SQM & build item table", key="calc_cols"):
        calc_df = build_items_from_columns(
            df=df,
            size_row=size_row,
            material_row=material_row,
            qty_annum_row=qty_annum_row,
            qty_run_row=qty_run_row,
            runs_pa_row=runs_pa_row,
            side_mode=side_mode,
            side_row=side_row,
            side_source_row=side_source_row,
            ds_synonyms=ds_synonyms,
            ss_synonyms=ss_synonyms,
        )
        st.session_state["calc_df"] = calc_df

# ---------- Show calculation results + Material group pricing ----------

if st.session_state["calc_df"] is not None:
    calc_df = st.session_state["calc_df"]

    st.subheader("Calculated SQM table (before pricing, numeric)")
    # Round sqm and qty to 2 decimals for clarity
    for col in ["Qty per annum", "Qty per run", "SQM per unit", "SQM per annum", "SQM per run"]:
        if col in calc_df.columns:
            calc_df[col] = calc_df[col].round(2)
    st.dataframe(calc_df)

    # ---------- Material groups + pricing presets ----------
    st.subheader("Material Groups & Pricing Presets")

    # Load saved preset (optional override of default) – only ONCE per uploaded file
    preset_file = st.file_uploader(
        "Load saved material groups preset (JSON, optional, overrides default ONCE)",
        type=["json"],
        key="group_preset_uploader",
    )
    if preset_file is not None and not st.session_state["preset_file_loaded_once"]:
        try:
            preset = json.load(preset_file)
            st.session_state["group_assignments"] = preset.get("group_assignments", {})
            st.session_state["group_prices"] = preset.get("group_prices", {})
            st.session_state["material_overrides"] = preset.get("material_overrides", {})
            st.session_state["preset_file_loaded_once"] = True
            st.success("Loaded group preset from uploaded file (will not auto-reload on each keystroke).")
        except Exception as e:
            st.error(f"Failed to load preset: {e}")

    # --------- Prep basic lists / dicts ---------
    materials = sorted(
        {m for m in calc_df["Material"].dropna().unique()} if "Material" in calc_df.columns else []
    )

    # Always pull latest from session_state
    group_assignments = st.session_state["group_assignments"]
    group_prices = st.session_state["group_prices"]
    group_volume_flags = st.session_state["group_volume_flags"]

    # ---------- STEP 1: Assign materials to groups (form, first) ----------
    st.markdown("**Step 1 – Assign materials to groups**")

    # Initialise widget state keys if missing
    if "existing_group_choice" not in st.session_state:
        st.session_state["existing_group_choice"] = "SelectExisting/None"
    if "group_name_input" not in st.session_state:
        st.session_state["group_name_input"] = ""
    if "assign_materials" not in st.session_state:
        st.session_state["assign_materials"] = []

    # If flagged, reset widget values, then clear flag
    if st.session_state.get("reset_assign_widgets", False):
        st.session_state["existing_group_choice"] = "SelectExisting/None"
        st.session_state["group_name_input"] = ""
        st.session_state["assign_materials"] = []
        st.session_state["reset_assign_widgets"] = False

    # Compute existing_groups & used_groups based on current assignments
    existing_groups = sorted({g for g in group_assignments.values() if g} | {g for g in group_prices.keys() if g})

    # Only show unassigned materials in the dropdown
    unassigned_materials = [m for m in materials if m not in group_assignments]

    with st.form("assign_group_form"):
        selected_materials = st.multiselect(
            "Select material(s) to assign to a group (only unassigned materials are shown)",
            options=unassigned_materials,
            key="assign_materials",
        )

        col_g1, col_g2 = st.columns(2)
        with col_g1:
            existing_group_choice = st.selectbox(
                "Pick existing group (optional)",
                options=["SelectExisting/None"] + existing_groups,
                index=0,
                key="existing_group_choice",
            )
        with col_g2:
            group_name_input = st.text_input(
                "Or type new group name",
                key="group_name_input",
            )

        submitted_assign = st.form_submit_button("Apply group to selected materials")

    if submitted_assign:
        manual = group_name_input.strip()
        if manual:
            group_name = manual
        elif existing_group_choice != "SelectExisting/None":
            group_name = existing_group_choice
        else:
            group_name = ""

        if not group_name:
            st.warning("Please type a new group name or pick an existing one.")
        elif not selected_materials:
            st.warning("Please select at least one material.")
        else:
            for m in selected_materials:
                group_assignments[m] = group_name
            st.session_state["group_assignments"] = group_assignments
            # Ensure volume flag has a default True for new groups
            if group_name not in group_volume_flags:
                group_volume_flags[group_name] = True
            st.session_state["group_volume_flags"] = group_volume_flags
            # Reset widgets on next rerun
            st.session_state["reset_assign_widgets"] = True
            st.success(f"Assigned group '{group_name}' to {len(selected_materials)} material(s).")

    # Refresh assignments after possible updates from the form
    group_assignments = st.session_state["group_assignments"]
    group_volume_flags = st.session_state["group_volume_flags"]

    # Recompute existing and used groups
    existing_groups = sorted({g for g in group_assignments.values() if g} | {g for g in group_prices.keys() if g})
    used_groups = set(group_assignments.values())
    used_default_groups = used_groups & DEFAULT_GROUP_NAMES

    # ---------- STEP 1b: Editable mapping table (after form) ----------
    mapping_df = pd.DataFrame({
        "Material": materials,
        "Group": [group_assignments.get(m, "") for m in materials]
    })

    st.markdown("**Current material → group mapping (editable)**")
    edited_mapping_df = st.data_editor(
        mapping_df,
        num_rows="fixed",
        use_container_width=True,
        disabled=["Material"],
        key="mapping_editor",
    )

    if st.button("Apply changes from mapping table"):
        new_assignments = {}
        for _, row in edited_mapping_df.iterrows():
            m = row["Material"]
            g = row["Group"]
            if isinstance(g, str):
                g = g.strip()
            if g:
                new_assignments[m] = g
        st.session_state["group_assignments"] = new_assignments
        group_assignments = new_assignments
        st.success("Updated group assignments from mapping table.")

        # Recompute groups after applying
        existing_groups = sorted({g for g in group_assignments.values() if g} | {g for g in group_prices.keys() if g})
        used_groups = set(group_assignments.values())
        used_default_groups = used_groups & DEFAULT_GROUP_NAMES

    # ---------- STEP 2: Set group base prices (per SQM, AUD) + volume toggle ----------
    st.markdown("**Step 2 – Set group BASE prices (per SQM, AUD) and volume pricing toggle**")

    all_groups = sorted(
        {g for g in group_assignments.values() if g} |
        {g for g in group_prices.keys() if g}
    )

    # Auto-hide groups not relevant for this campaign (not used in this sheet)
    not_relevant_groups = sorted(set(all_groups) - used_groups)

    extra_groups_to_show = st.multiselect(
        "Also show these groups (not used in this campaign):",
        options=not_relevant_groups,
        default=st.session_state.get("extra_groups_to_show", []),
        key="extra_groups_to_show_selector",
    )
    st.session_state["extra_groups_to_show"] = extra_groups_to_show

    visible_groups = sorted(list(used_groups | set(extra_groups_to_show)))

    new_group_prices = {}
    new_group_volume_flags = {}

    if not all_groups:
        st.info("No groups yet. Assign at least one material to a group in Step 1.")
    else:
        st.caption(
            "Colour legend: "
            "<span style='color:green;'>Green = default group (not used in this sheet)</span>, "
            "<span style='color:#ff8800;'>Orange = default group used in this sheet</span>, "
            "<span style='color:red;'>Red = new group</span>. "
            "Checkbox: whether this group uses the global volume tiers or just flat BASE price.",
            unsafe_allow_html=True,
        )
        # Show base price inputs in a 4-column grid, only for visible groups
        cols_per_row = 4
        for i, g in enumerate(visible_groups):
            if i % cols_per_row == 0:
                cols = st.columns(cols_per_row)
            col = cols[i % cols_per_row]

            existing_price = group_prices.get(g)
            if existing_price is None or (isinstance(existing_price, float) and math.isnan(existing_price)):
                initial_value = 0.0
            else:
                try:
                    initial_value = float(existing_price)
                except Exception:
                    initial_value = 0.0

            # Colour caption:
            # - orange for default groups actually used in this sheet
            # - green for other default groups
            # - red for new groups
            is_default_group = g in DEFAULT_GROUP_NAMES
            is_used_default = g in used_default_groups
            if is_default_group and is_used_default:
                caption_color = "#ff8800"  # bright orange
            elif is_default_group:
                caption_color = "green"
            else:
                caption_color = "red"

            with col:
                st.markdown(
                    f"<div style='color:{caption_color}; font-weight:600; font-size:0.85rem;'>"
                    f"BASE price per SQM (AUD) for group '{g}'"
                    f"</div>",
                    unsafe_allow_html=True,
                )
                price = st.number_input(
                    "",
                    min_value=0.0,
                    max_value=1_000_000.0,
                    step=0.01,
                    format="%.2f",
                    value=initial_value,
                    key=f"group_price_input_{g}",
                )
                use_volume = st.checkbox(
                    "Use volume tiers",
                    value=group_volume_flags.get(g, True),
                    key=f"group_use_volume_{g}",
                    help="If unticked, this group always uses the flat BASE price (no tier discount).",
                )

            if price == 0.0 and (existing_price is None or (isinstance(existing_price, float) and math.isnan(existing_price))):
                new_group_prices[g] = np.nan
            else:
                new_group_prices[g] = price

            new_group_volume_flags[g] = use_volume

        # Keep prices & volume flags for groups not shown in the UI unchanged
        for g in all_groups:
            if g not in visible_groups:
                if g in group_prices:
                    new_group_prices[g] = group_prices[g]
                else:
                    new_group_prices[g] = np.nan
                if g in group_volume_flags:
                    new_group_volume_flags[g] = group_volume_flags[g]
                else:
                    new_group_volume_flags[g] = True

        st.session_state["group_prices"] = new_group_prices
        st.session_state["group_volume_flags"] = new_group_volume_flags
        group_prices = new_group_prices
        group_volume_flags = new_group_volume_flags

    # Build preset data from current state (overrides kept but not used in pricing)
    preset_data = {
        "group_assignments": st.session_state["group_assignments"],
        "group_prices": st.session_state["group_prices"],
        "group_volume_flags": st.session_state["group_volume_flags"],
        "material_overrides": st.session_state["material_overrides"],
    }

    # ---------- Session snapshots (save this workbook's working state) ----------
    st.subheader("Session snapshots (save this workbook's pricing setup)")
    col_s1, col_s2 = st.columns([2, 1])

    with col_s1:
        snapshot_name = st.text_input(
            "Snapshot name (e.g. 'BP Tender Jan', 'FootLocker Xmas')",
            key="snapshot_name",
        )
    with col_s2:
        if st.button("Save snapshot"):
            name = snapshot_name.strip()
            if not name:
                st.warning("Please enter a snapshot name before saving.")
            else:
                snapshot_payload = {
                    "group_assignments": st.session_state["group_assignments"],
                    "group_prices": st.session_state["group_prices"],
                    "group_volume_flags": st.session_state["group_volume_flags"],
                    "material_overrides": st.session_state["material_overrides"],
                    "extra_groups_to_show": st.session_state["extra_groups_to_show"],
                    "ds_syn_input": st.session_state["ds_syn_input"],
                    "ss_syn_input": st.session_state["ss_syn_input"],
                    "double_sided_loading_percent": st.session_state["double_sided_loading_percent"],
                    "layout_type_choice": st.session_state["layout_type_choice"],
                    "hidden_cols_letters": st.session_state["hidden_cols_letters"],
                    "hidden_rows_numbers": st.session_state["hidden_rows_numbers"],
                    "tier_count": st.session_state["tier_count"],
                    "tier_thresholds": st.session_state["tier_thresholds"],
                    "tier_discounts": st.session_state["tier_discounts"],
                }
                st.session_state["session_snapshots"][name] = snapshot_payload
                st.success(f"Saved snapshot '{name}'. You can load it later in this browser session.")

    if st.session_state["session_snapshots"]:
        col_l1, col_l2 = st.columns([2, 1])
        with col_l1:
            snapshot_options = ["(none)"] + list(st.session_state["session_snapshots"].keys())
            snapshot_to_load = st.selectbox(
                "Load existing snapshot",
                options=snapshot_options,
                key="snapshot_to_load",
            )
        with col_l2:
            if st.button("Load snapshot"):
                if snapshot_to_load != "(none)":
                    snap = st.session_state["session_snapshots"][snapshot_to_load]
                    st.session_state["group_assignments"] = snap.get("group_assignments", {})
                    st.session_state["group_prices"] = snap.get("group_prices", {})
                    st.session_state["group_volume_flags"] = snap.get("group_volume_flags", {})
                    st.session_state["material_overrides"] = snap.get("material_overrides", {})
                    st.session_state["extra_groups_to_show"] = snap.get("extra_groups_to_show", [])
                    # restore config
                    if "ds_syn_input" in snap:
                        st.session_state["ds_syn_input"] = snap["ds_syn_input"]
                    if "ss_syn_input" in snap:
                        st.session_state["ss_syn_input"] = snap["ss_syn_input"]
                    if "double_sided_loading_percent" in snap:
                        st.session_state["double_sided_loading_percent"] = snap["double_sided_loading_percent"]
                    if "layout_type_choice" in snap:
                        st.session_state["layout_type_choice"] = snap["layout_type_choice"]
                    if "hidden_cols_letters" in snap:
                        st.session_state["hidden_cols_letters"] = snap["hidden_cols_letters"]
                    if "hidden_rows_numbers" in snap:
                        st.session_state["hidden_rows_numbers"] = snap["hidden_rows_numbers"]
                    if "tier_count" in snap:
                        st.session_state["tier_count"] = snap["tier_count"]
                    if "tier_thresholds" in snap:
                        st.session_state["tier_thresholds"] = snap["tier_thresholds"]
                    if "tier_discounts" in snap:
                        st.session_state["tier_discounts"] = snap["tier_discounts"]
                    st.success(f"Loaded snapshot '{snapshot_to_load}'. Rerun mappings if you've changed Excel.")
                else:
                    st.info("Select a snapshot to load.")

    st.markdown("**Preset saving options**")
    col_left, col_right = st.columns(2)

    with col_left:
        save_mode = st.radio(
            "How should these pricing changes be saved?",
            [
                "Only for this session (do not update JSON on server)",
                "Update default JSON on server (material_groups_default.json)",
            ],
            index=0,
            help=(
                "If you choose to update the default JSON, the app will try to overwrite "
                "'material_groups_default.json' in the current environment. On Streamlit Cloud "
                "this may not persist across deployments; use the download option to commit to GitHub."
            ),
        )

        if st.button("Apply save option now"):
            if save_mode.startswith("Update default JSON"):
                try:
                    with open("material_groups_default.json", "w", encoding="utf-8") as f:
                        json.dump(preset_data, f, indent=2)
                    st.success("Updated material_groups_default.json on server.")
                except Exception as e:
                    st.error(f"Could not update default JSON on server: {e}")
            else:
                st.info("Changes kept only in this session (JSON file not modified).")

    with col_right:
        preset_bytes = BytesIO(json.dumps(preset_data, indent=2).encode("utf-8"))
        st.download_button(
            "Download current material group preset (JSON)",
            data=preset_bytes,
            file_name="material_groups_preset.json",
            mime="application/json",
            help=(
                "Download the latest preset and commit it to your Git repo if you want it "
                "to be the new default next time you deploy."
            ),
        )

    # ---------- Apply pricing ----------

    st.subheader("Apply grouped pricing")

    group_assignment_map = st.session_state["group_assignments"]
    group_price_map = st.session_state["group_prices"]
    group_volume_flags = st.session_state["group_volume_flags"]

    calc_with_price = calc_df.copy()

    # Resolve base price per SQM (AUD) for each row from group only
    def resolve_base_price(material):
        group = group_assignment_map.get(material)
        if group:
            gp = group_price_map.get(group)
            if gp is not None and not pd.isna(gp):
                return gp
        return np.nan

    calc_with_price["Base Price per SQM (AUD)"] = calc_with_price["Material"].apply(
        resolve_base_price
    )

# ---------- Volume-based price breaks (global, by SQM per annum) ----------
st.markdown("**Volume-based price breaks (by SQM per annum, global)**")
st.caption(
    "Configure SQM tiers and % adjustments vs the BASE group price. "
    "Each group can opt in/out. Example: "
    "Tier1 up to 250sqm = 0%, Tier2 up to 500sqm = -1%, Tier3 up to 1000sqm = -2%. "
    "Any volume above the highest tier uses the last tier's %."
)

# Let user choose how many tiers (ranges)
tier_count = st.number_input(
    "Number of SQM tiers (ranges)",
    min_value=1,
    max_value=10,
    step=1,
    value=st.session_state["tier_count"],
    key="tier_count_input",
    help="This controls how many SQM ranges you want (e.g. 3 = up to 250, 500, 1000).",
)
st.session_state["tier_count"] = int(tier_count)

# Make sure stored lists match tier_count length
while len(st.session_state["tier_thresholds"]) < tier_count:
    last_thr = st.session_state["tier_thresholds"][-1] if st.session_state["tier_thresholds"] else 250.0
    st.session_state["tier_thresholds"].append(last_thr)
while len(st.session_state["tier_discounts"]) < tier_count:
    st.session_state["tier_discounts"].append(0.0)
if len(st.session_state["tier_thresholds"]) > tier_count:
    st.session_state["tier_thresholds"] = st.session_state["tier_thresholds"][: tier_count]
if len(st.session_state["tier_discounts"]) > tier_count:
    st.session_state["tier_discounts"] = st.session_state["tier_discounts"][: tier_count]

# Build an interactive table for tiers
tier_rows = []
for i in range(int(tier_count)):
    tier_rows.append(
        {
            "Tier": i + 1,
            "Max SQM": float(st.session_state["tier_thresholds"][i]),
            "% vs base": float(st.session_state["tier_discounts"][i]),
        }
    )

tier_df = pd.DataFrame(tier_rows)

st.write("Configure tiers (edit cells below):")
edited_tier_df = st.data_editor(
    tier_df,
    num_rows="fixed",
    use_container_width=True,
    disabled=["Tier"],
    key="tier_editor",
)

# Push edited values back into session_state
tier_thresholds = []
tier_discounts = []
for _, r in edited_tier_df.iterrows():
    # force non-negative thresholds
    max_sqm = float(r["Max SQM"]) if not pd.isna(r["Max SQM"]) else 0.0
    if max_sqm < 0:
        max_sqm = 0.0
    tier_thresholds.append(max_sqm)
    tier_discounts.append(float(r["% vs base"]) if not pd.isna(r["% vs base"]) else 0.0)

st.session_state["tier_thresholds"] = tier_thresholds
st.session_state["tier_discounts"] = tier_discounts

# Sort tiers by threshold so ranges are consistent
tiers_sorted = sorted(zip(tier_thresholds, tier_discounts), key=lambda x: x[0])
sorted_thresholds = [t[0] for t in tiers_sorted]
sorted_discounts = [t[1] for t in tiers_sorted]

def pick_discount_for_sqm(sqm):
    if pd.isna(sqm) or not sorted_thresholds:
        return 0.0
    for T, D in zip(sorted_thresholds, sorted_discounts):
        if sqm <= T:
            return D
    return sorted_discounts[-1]  # use last tier for > max

# Determine volume discount per row based on SQM per annum (fallback to SQM per run)
volume_discounts = []
tier_prices = []
for _, row in calc_with_price.iterrows():
    base = row["Base Price per SQM (AUD)"]
    material = row.get("Material")
    if pd.isna(base):
        volume_discounts.append(0.0)
        tier_prices.append(np.nan)
        continue

    group = group_assignment_map.get(material)
    use_tiers = group_volume_flags.get(group, True)

    if not use_tiers:
        disc = 0.0
        tier_price = base
    else:
        sqm_metric = row.get("SQM per annum")
        if pd.isna(sqm_metric):
            sqm_metric = row.get("SQM per run")
        disc = pick_discount_for_sqm(sqm_metric)
        tier_price = base * (1.0 + disc / 100.0)

    volume_discounts.append(disc)
    tier_prices.append(tier_price)

calc_with_price["Volume Discount %"] = volume_discounts
calc_with_price["Tier Price per SQM (AUD)"] = tier_prices

    # Apply DS loading to get effective price (AUD)
    ds_factor = 1.0 + st.session_state["double_sided_loading_percent"] / 100.0
    calc_with_price["Effective Price per SQM (AUD)"] = calc_with_price.apply(
        lambda r: r["Tier Price per SQM (AUD)"] * ds_factor
        if r.get("Side") == "DS"
        else r["Tier Price per SQM (AUD)"],
        axis=1,
    )

    # Price calculations in AUD
    calc_with_price["Price per unit (AUD)"] = (
        calc_with_price["SQM per unit"] * calc_with_price["Effective Price per SQM (AUD)"]
    )
    calc_with_price["Price per annum (AUD)"] = (
        calc_with_price["SQM per annum"] * calc_with_price["Effective Price per SQM (AUD)"]
    )
    calc_with_price["Price per run (AUD)"] = (
        calc_with_price["SQM per run"] * calc_with_price["Effective Price per SQM (AUD)"]
    )

    # Round all numeric price-related columns to 2 decimals
    price_cols = [c for c in calc_with_price.columns if "Price" in c]
    for col in price_cols:
        calc_with_price[col] = calc_with_price[col].round(2)
    if "Volume Discount %" in calc_with_price.columns:
        calc_with_price["Volume Discount %"] = calc_with_price["Volume Discount %"].round(2)

    # ---------- Group-level sqm + price summary ----------
    if "Material" in calc_with_price.columns:
        calc_with_group = calc_with_price.copy()
        calc_with_group["Group"] = calc_with_group["Material"].map(group_assignment_map).fillna("")
        grouped_rows = calc_with_group[calc_with_group["Group"] != ""]
        if not grouped_rows.empty:
            # Sum sqm and price per group
            agg_dict = {}
            for c in ["SQM per unit", "SQM per annum", "SQM per run"]:
                if c in grouped_rows.columns:
                    agg_dict[c] = "sum"
            for c in ["Price per annum (AUD)", "Price per run (AUD)"]:
                if c in grouped_rows.columns:
                    agg_dict[c] = "sum"

            group_summary_full = grouped_rows.groupby("Group").agg(agg_dict).reset_index()

            # Attach base group price per SQM (AUD)
            group_summary_full["Base Group Price per SQM (AUD)"] = (
                group_summary_full["Group"].map(group_price_map).round(2)
            )

            # Round sqm and price columns
            for c in group_summary_full.columns:
                if c != "Group":
                    group_summary_full[c] = group_summary_full[c].round(2)

            st.subheader("Square metres & prices by Group (including per run & per annum)")
            st.dataframe(group_summary_full, use_container_width=True)

            # Download
            group_buf = BytesIO()
            group_summary_full.to_excel(group_buf, index=False, sheet_name="GROUP_SQM_PRICE")
            group_buf.seek(0)
            st.download_button(
                "Download group sqm & price summary (Excel)",
                data=group_buf,
                file_name="group_sqm_price_summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # Build a display copy with $ sign for price columns
    display_df = calc_with_price.copy()
    for col in price_cols:
        display_df[col] = display_df[col].apply(
            lambda x: "" if pd.isna(x) else f"${x:,.2f}"
        )

    st.subheader("Final calculation table (with grouped pricing, formatted)")
    st.dataframe(display_df)

    # Download calculated table (numeric, rounded) as separate CALC workbook
    out_calc = BytesIO()
    calc_with_price.to_excel(out_calc, index=False, sheet_name="CALC")
    out_calc.seek(0)

    st.download_button(
        "Download SQM & pricing table (CALC.xlsx)",
        data=out_calc,
        file_name="sqm_pricing_calc.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    # ---------- Write results back into original workbook ----------

    st.subheader("Write SQM & pricing back into original workbook")

    # Helper to normalise column letters typed by the user
    def normalise_col_letters(s: str):
        if not s:
            return None
        s = s.strip().upper()
        if not re.fullmatch(r"[A-Z]+", s):
            return None
        return s

    if layout_type == "Items are in rows (BP-style)":
        st.caption("Type the exact Excel column letters where you want SQM and prices to go (e.g. N, Q, AA). Leave blank to skip a field.")

        sqm_annum_col_str = st.text_input(
            "Excel column for SQM per annum (optional, e.g. N or AA)",
            value="",
            key="out_sqm_annum_col_str",
        )
        sqm_run_col_str = st.text_input(
            "Excel column for SQM per run (optional, e.g. O or AB)",
            value="",
            key="out_sqm_run_col_str",
        )
        price_annum_col_str = st.text_input(
            "Excel column for Price per annum (AUD) (optional, e.g. P or AC)",
            value="",
            key="out_price_annum_col_str",
        )
        price_run_col_str = st.text_input(
            "Excel column for Price per run (AUD) (optional, e.g. Q or AD)",
            value="",
            key="out_price_run_col_str",
        )

        sqm_annum_col_letter = normalise_col_letters(sqm_annum_col_str)
        sqm_run_col_letter = normalise_col_letters(sqm_run_col_str)
        price_annum_col_letter = normalise_col_letters(price_annum_col_str)
        price_run_col_letter = normalise_col_letters(price_run_col_str)

        if st.button("Build original workbook with SQM & prices filled (rows layout)"):
            if not any([sqm_annum_col_letter, sqm_run_col_letter, price_annum_col_letter, price_run_col_letter]):
                st.warning("Please type at least one output column (e.g. N, Q, AA).")
            else:
                wb2 = load_workbook(BytesIO(file_bytes))
                ws2 = wb2[sheet_name]

                for _, r in calc_with_price.iterrows():
                    src_row = r.get("Source Row")
                    if pd.isna(src_row):
                        continue
                    excel_row = int(src_row)

                    if sqm_annum_col_letter:
                        val = r.get("SQM per annum")
                        if val is not None and not pd.isna(val):
                            ws2[f"{sqm_annum_col_letter}{excel_row}"] = float(val)
                    if sqm_run_col_letter:
                        val = r.get("SQM per run")
                        if val is not None and not pd.isna(val):
                            ws2[f"{sqm_run_col_letter}{excel_row}"] = float(val)
                    if price_annum_col_letter:
                        val = r.get("Price per annum (AUD)")
                        if val is not None and not pd.isna(val):
                            ws2[f"{price_annum_col_letter}{excel_row}"] = float(val)
                    if price_run_col_letter:
                        val = r.get("Price per run (AUD)")
                        if val is not None and not pd.isna(val):
                            ws2[f"{price_run_col_letter}{excel_row}"] = float(val)

                out_buf2 = BytesIO()
                wb2.save(out_buf2)
                out_buf2.seek(0)

                st.download_button(
                    "Download ORIGINAL workbook with SQM & prices filled (rows)",
                    data=out_buf2,
                    file_name=f"{sheet_name}_with_pricing_rows.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    elif layout_type == "Items are in columns (Foot Locker-style)":
        st.caption("Type the exact Excel row numbers where you want SQM and prices to go across each material column (e.g. 150, 200). Leave blank to skip a field.")

        sqm_annum_row_str = st.text_input(
            "Row for SQM per annum (across columns, optional, e.g. 120)",
            value="",
            key="out_sqm_annum_row_str",
        )
        sqm_run_row_str = st.text_input(
            "Row for SQM per run (across columns, optional, e.g. 121)",
            value="",
            key="out_sqm_run_row_str",
        )
        price_annum_row_str = st.text_input(
            "Row for Price per annum (AUD) (across columns, optional, e.g. 122)",
            value="",
            key="out_price_annum_row_str",
        )
        price_run_row_str = st.text_input(
            "Row for Price per run (AUD) (across columns, optional, e.g. 123)",
            value="",
            key="out_price_run_row_str",
        )

        def parse_row_num(s: str):
            if not s:
                return None
            s = s.strip()
            if not s.isdigit():
                return None
            n = int(s)
            return n if n > 0 else None

        sqm_annum_row = parse_row_num(sqm_annum_row_str)
        sqm_run_row = parse_row_num(sqm_run_row_str)
        price_annum_row = parse_row_num(price_annum_row_str)
        price_run_row = parse_row_num(price_run_row_str)

        if st.button("Build original workbook with SQM & prices filled (columns layout)"):
            if not any([sqm_annum_row, sqm_run_row, price_annum_row, price_run_row]):
                st.warning("Please type at least one output row (e.g. 120, 150, 200).")
            else:
                wb2 = load_workbook(BytesIO(file_bytes))
                ws2 = wb2[sheet_name]

                for _, r in calc_with_price.iterrows():
                    src_col = r.get("Source Column")
                    if src_col is None or (isinstance(src_col, float) and pd.isna(src_col)):
                        continue
                    col_letter = str(src_col)

                    if sqm_annum_row is not None:
                        val = r.get("SQM per annum")
                        if val is not None and not pd.isna(val):
                            ws2[f"{col_letter}{sqm_annum_row}"] = float(val)
                    if sqm_run_row is not None:
                        val = r.get("SQM per run")
                        if val is not None and not pd.isna(val):
                            ws2[f"{col_letter}{sqm_run_row}"] = float(val)
                    if price_annum_row is not None:
                        val = r.get("Price per annum (AUD)")
                        if val is not None and not pd.isna(val):
                            ws2[f"{col_letter}{price_annum_row}"] = float(val)
                    if price_run_row is not None:
                        val = r.get("Price per run (AUD)")
                        if val is not None and not pd.isna(val):
                            ws2[f"{col_letter}{price_run_row}"] = float(val)

                out_buf2 = BytesIO()
                wb2.save(out_buf2)
                out_buf2.seek(0)

                st.download_button(
                    "Download ORIGINAL workbook with SQM & prices filled (columns)",
                    data=out_buf2,
                    file_name=f"{sheet_name}_with_pricing_columns.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
